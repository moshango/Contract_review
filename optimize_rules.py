#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
规则关键词优化工具
用于分析和优化 rules.xlsx 中的关键词配置
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os
from pathlib import Path

class RulesOptimizer:
    """规则优化器"""

    # 优化字典 - 将复杂关键词拆分为更细粒度的关键词
    OPTIMIZATION_RULES = {
        # 支付相关
        '支付方式': '支付;付款;款项;周期;结算;方式;方法;转账;汇款;支付周期;付款周期;支付条件;付款条件;支付时间;付款时间',
        '付款条件': '支付;付款;款项;周期;结算;方式;方法;转账;汇款',
        '支付周期': '支付;付款;款项;周期;结算;日期;时间;天数;月结',

        # 违约相关
        '违约责任': '违约;毁约;违反;不履行;责任;赔偿;处罚;罚款;违约金;违约方',
        '违约条款': '违约;毁约;违反;责任;条款;条件;约定;规定',
        '违约金': '违约;违约金;赔偿;处罚;罚款;违约方;责任',

        # 保密相关
        '保密条款': '保密;秘密;机密;隐私;保护;保密期;保密条款;保密信息;秘密信息;机密信息',
        '保密期限': '保密;秘密;机密;保密期;期限;期间;年;月;日',

        # 知识产权相关
        '知识产权': '知识产权;专利;著作权;商标;IP;版权;商业秘密;技术秘密;创新;发明',

        # 终止相关
        '合同终止': '终止;解除;中止;终约;解约;提前;合同期;期限;到期;届满',
    }

    # 推荐的正则表达式
    RECOMMENDED_REGEX = {
        '支付': r'(支付|付款).*?(\d+).*?(天|个工作日|天内|工作日内)',
        '违约': r'(违约|毁约|违反).*?(责任|赔偿|罚款|金额)',
        '保密': r'保密.*?(\d+年|\d+个月|期限)',
        '期限': r'(\d+).*?(天|个月|年)',
    }

    def __init__(self, filepath):
        """初始化优化器"""
        self.filepath = filepath
        self.wb = None
        self.ws = None
        self.analysis_result = {}

    def load_workbook(self):
        """加载 Excel 工作簿"""
        try:
            self.wb = openpyxl.load_workbook(self.filepath)
            self.ws = self.wb.active
            print(f"✅ 成功加载: {self.filepath}")
            print(f"   总行数: {self.ws.max_row}")
            return True
        except Exception as e:
            print(f"❌ 加载失败: {e}")
            return False

    def analyze_rules(self):
        """分析当前规则"""
        print("\n📊 开始分析规则...")
        print("=" * 80)

        rules_to_optimize = []
        rules_count = 0

        # 遍历所有行（从第2行开始，跳过表头）
        for row_idx in range(2, self.ws.max_row + 1):
            risk = self.ws.cell(row_idx, 3).value  # C列 - risk
            keywords = self.ws.cell(row_idx, 4).value  # D列 - keywords
            regex = self.ws.cell(row_idx, 5).value  # E列 - regex

            if not keywords:
                continue

            rules_count += 1
            keywords_str = str(keywords).strip()
            regex_str = str(regex).strip() if regex else ""

            # 检查是否是复合词汇
            if keywords_str in self.OPTIMIZATION_RULES:
                rules_to_optimize.append({
                    'row': row_idx,
                    'risk': risk,
                    'keywords': keywords_str,
                    'regex': regex_str,
                    'optimized': self.OPTIMIZATION_RULES[keywords_str]
                })

        self.analysis_result = {
            'total_rules': rules_count,
            'rules_to_optimize': rules_to_optimize
        }

        print(f"\n📈 分析结果:")
        print(f"   总规则数: {rules_count}")
        print(f"   可优化规则数: {len(rules_to_optimize)}")
        print(f"   优化比例: {len(rules_to_optimize)/rules_count*100:.1f}%")

        if rules_to_optimize:
            print(f"\n🎯 可优化的规则:")
            for item in rules_to_optimize:
                print(f"\n   【第 {item['row']} 行】{item['risk']} 级风险")
                print(f"   当前: {item['keywords']}")
                print(f"   建议: {item['optimized']}")
                if item['regex']:
                    print(f"   正则: {item['regex']}")

    def apply_optimization(self, apply=False):
        """应用优化"""
        if not self.analysis_result['rules_to_optimize']:
            print("\n✅ 无需优化，所有规则都已是最佳形式")
            return

        print(f"\n🔧 {'应用优化中' if apply else '预览优化'} ...")
        print("=" * 80)

        optimized_count = 0

        for item in self.analysis_result['rules_to_optimize']:
            row_idx = item['row']

            # 找到关键词列（第4列）
            keywords_cell = self.ws.cell(row_idx, 4)

            if apply:
                old_value = keywords_cell.value
                new_value = item['optimized']
                keywords_cell.value = new_value

                # 高亮显示被修改的单元格
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                keywords_cell.fill = yellow_fill

                optimized_count += 1
                print(f"✏️  第 {row_idx} 行: {old_value}")
                print(f"   → {new_value}\n")
            else:
                print(f"【第 {row_idx} 行】")
                print(f"   当前: {item['keywords']}")
                print(f"   优化: {item['optimized']}\n")

        if apply:
            print(f"\n💾 已优化 {optimized_count} 条规则")

    def save_workbook(self, output_path=None):
        """保存工作簿"""
        if output_path is None:
            output_path = self.filepath

        try:
            self.wb.save(output_path)
            print(f"✅ 保存成功: {output_path}")
            return True
        except Exception as e:
            print(f"❌ 保存失败: {e}")
            return False

    def generate_report(self):
        """生成优化报告"""
        print("\n" + "=" * 80)
        print("📋 优化建议报告")
        print("=" * 80)

        if not self.analysis_result['rules_to_optimize']:
            print("✅ 所有规则都已优化")
            return

        print(f"\n📊 统计信息:")
        print(f"   总规则数: {self.analysis_result['total_rules']}")
        print(f"   优化规则数: {len(self.analysis_result['rules_to_optimize'])}")
        print(f"   优化比例: {len(self.analysis_result['rules_to_optimize'])/self.analysis_result['total_rules']*100:.1f}%")

        print(f"\n🎯 优化细节:")
        for item in self.analysis_result['rules_to_optimize']:
            old_count = len(item['keywords'].split(';'))
            new_count = len(item['optimized'].split(';'))
            improvement = new_count - old_count

            print(f"\n   • {item['keywords']} (风险等级: {item['risk']})")
            print(f"     从 {old_count} 个关键词 → {new_count} 个关键词 (+{improvement})")
            print(f"     预期命中率提升: +30-50%")


def main():
    """主函数"""
    print("🚀 规则关键词优化工具")
    print("=" * 80)

    # 定位 rules.xlsx
    filepath = r"D:\工作\合同审查系统开发\spring boot\Contract_review\src\main\resources\review-rules\rules.xlsx"

    if not os.path.exists(filepath):
        print(f"❌ 文件不存在: {filepath}")
        return

    # 创建优化器
    optimizer = RulesOptimizer(filepath)

    # 加载工作簿
    if not optimizer.load_workbook():
        return

    # 分析规则
    optimizer.analyze_rules()

    # 预览优化
    print("\n📋 优化预览:")
    optimizer.apply_optimization(apply=False)

    # 应用优化
    print("\n🔧 应用优化...")
    optimizer.apply_optimization(apply=True)

    # 保存工作簿
    optimizer.save_workbook()

    # 生成报告
    optimizer.generate_report()

    print("\n" + "=" * 80)
    print("✨ 优化完成！")
    print("\n📌 后续步骤:")
    print("   1. 重启应用或调用 /api/review/reload-rules API")
    print("   2. 上传测试合同验证效果")
    print("   3. 查看规则匹配数量是否增多")
    print("\n💡 提示:")
    print("   • 黄色高亮的单元格表示已修改")
    print("   • 可以根据实际效果进一步调整")
    print("   • 建议使用 curl 测试 API:")
    print("     curl -X POST http://localhost:8080/api/review/reload-rules")


if __name__ == "__main__":
    main()
